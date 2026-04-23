"""pxts I/O layer — read_csv, write_csv, read_bdh, read_mb, and read_xlsx.

Public API:
    read_csv(path, *, tz=None, date_format=None) -> pd.DataFrame
    write_csv(df, path, *, date_format=None) -> None
    read_bdh(tickers, start, field='PX_LAST', end=None) -> pd.DataFrame
    read_mb(series) -> pd.DataFrame
    read_xlsx(path, *, tz=None, date_format=None, datetime_col='A',
              colname_row=1, values_ref=None, sheet=None) -> pd.DataFrame

Internal:
    _detect_date_format(sample) -> tuple[str, bool]
"""
from __future__ import annotations

import datetime
import re
import warnings
from pathlib import Path
from typing import Union

import pandas as pd

from pxts.core import validate_ts

# ---------------------------------------------------------------------------
# Compiled regex patterns for date format detection
# ---------------------------------------------------------------------------

# Matches ISO 8601: YYYY-MM-DD or YYYY/MM/DD
_ISO_RE = re.compile(r"^\d{4}[-/]\d{2}[-/]\d{2}")

# Matches slash-delimited dates: D/M/YYYY or DD/MM/YYYY or M/D/YYYY etc.
_SLASH_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})")


def _detect_date_format(sample: str) -> tuple[str, bool]:
    """Detect date format from a sample date string.

    Returns a tuple of (format_string, dayfirst) where:
    - format_string is one of "ISO8601", "%d/%m/%Y", "%m/%d/%Y", or "mixed"
    - dayfirst is True for DD/MM/YYYY format, False otherwise

    Rules:
    - ISO match (YYYY-MM-DD or YYYY/MM/DD): return ("ISO8601", False)
    - Slash match: if first part > 12 -> DD/MM/YYYY ("%d/%m/%Y", True)
                   if second part > 12 -> MM/DD/YYYY ("%m/%d/%Y", False)
                   ambiguous (both <= 12) -> default to British ("%d/%m/%Y", True)
    - Fallback: ("mixed", False)
    """
    sample = sample.strip()

    if _ISO_RE.match(sample):
        return ("ISO8601", False)

    slash_match = _SLASH_RE.match(sample)
    if slash_match:
        first = int(slash_match.group(1))
        second = int(slash_match.group(2))
        if first > 12:
            # Day must be first — UK format DD/MM/YYYY
            return ("%d/%m/%Y", True)
        elif second > 12:
            # Month first — US format MM/DD/YYYY
            return ("%m/%d/%Y", False)
        else:
            # Ambiguous — both parts <= 12 — default to British DD/MM/YYYY
            warnings.warn(
                f"pxts: ambiguous date '{sample}' — assumed DD/MM/YYYY (British). "
                f"Pass date_format='%m/%d/%Y' to override.",
                UserWarning,
                stacklevel=3,  # _detect_date_format -> read_csv -> user call site
            )
            return ("%d/%m/%Y", True)

    return ("mixed", False)


def read_csv(
    path: Union[str, Path],
    *,
    tz: str | None = None,
    date_format: str | None = None,
) -> pd.DataFrame:
    """Read a CSV file into a DataFrame with a DatetimeIndex.

    The first column is treated as the datetime index. The first row is
    column headers. Date format is auto-detected from the first data row —
    supports ISO 8601 (YYYY-MM-DD), US (MM/DD/YYYY), and UK (DD/MM/YYYY).

    Parameters
    ----------
    path : str or Path
        Path to the CSV file.
    tz : str or None
        If provided, localize (or convert) the index to this timezone.
        Example: tz='US/Eastern'
    date_format : str or None
        Explicit date format string (e.g. '%Y-%m-%d'). If None, auto-detects.

    Returns
    -------
    pd.DataFrame
        DataFrame with a DatetimeIndex as the index.
    """
    path = Path(path)

    if date_format is not None:
        # Caller provided explicit format — skip detection
        detected_format = date_format
        dayfirst = False
    else:
        # Two-pass: peek first row to detect format
        peek = pd.read_csv(path, nrows=1, header=0)
        sample_str = str(peek.iloc[0, 0])
        detected_format, dayfirst = _detect_date_format(sample_str)

    # Full read with index in first column
    df = pd.read_csv(path, index_col=0, header=0)

    # Save index name before conversion (will be lost during to_datetime)
    index_name = df.index.name

    # Convert string index to DatetimeIndex
    if detected_format == "ISO8601":
        df.index = pd.to_datetime(df.index, format="ISO8601")
    elif detected_format == "mixed":
        df.index = pd.to_datetime(df.index, format="mixed", dayfirst=dayfirst)
    else:
        df.index = pd.to_datetime(df.index, format=detected_format, dayfirst=dayfirst)

    # Restore index name
    df.index.name = index_name

    # Apply timezone if requested
    if tz is not None:
        if df.index.tz is None:
            df.index = df.index.tz_localize(tz)
        else:
            df.index = df.index.tz_convert(tz)

    return df


def write_csv(
    df: pd.DataFrame,
    path: Union[str, Path],
    *,
    date_format: str | None = None,
) -> None:
    """Write a DatetimeIndex DataFrame to a CSV file.

    The datetime index is written as the first column. The first row is
    column headers. Uses ISO 8601 with UTC offset by default for safe
    round-trips.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame with a DatetimeIndex. Raises pxtsValidationError otherwise.
    path : str or Path
        Destination CSV path.
    date_format : str or None
        Explicit date format string. Defaults to '%Y-%m-%dT%H:%M:%S%z'
        (ISO 8601 with offset — safe for round-trips).

    Raises
    ------
    pxtsValidationError
        If df does not have a DatetimeIndex.
    """
    validate_ts(df)

    if date_format is None:
        date_format = "%Y-%m-%dT%H:%M:%S%z"

    df.to_csv(path, index=True, date_format=date_format)


def read_bdh(
    tickers,
    start="2000-01-01",
    field: str = "PX_LAST",
    end=None,
) -> pd.DataFrame:
    """Fetch Bloomberg BDH historical time series data.

    Opens a Bloomberg connection, fetches the requested tickers over the
    given date range, and returns a validated wide-format DataFrame with
    a DatetimeIndex and one column per ticker.

    Requires pdblp (optional dependency). Install with:
        pip install pxts[bloomberg]

    Parameters
    ----------
    tickers : list of str, str, or dict
        Bloomberg ticker strings, e.g. ['AAPL US Equity', 'MSFT US Equity'].
        If a dict is provided, keys are the desired output column names and
        values are the Bloomberg tickers, e.g.
        {'Apple': 'AAPL US Equity', 'Microsoft': 'MSFT US Equity'}.
    start : str, datetime, or pd.Timestamp
        Start date (inclusive). Defaults to '2000-01-01'. Converted to 'YYYYMMDD' string internally.
    field : str
        Bloomberg field name. Defaults to 'PX_LAST'.
    end : str, datetime, pd.Timestamp, or None
        End date (inclusive). Defaults to today when None.

    Returns
    -------
    pd.DataFrame
        Wide-format DataFrame with DatetimeIndex (rows = dates,
        columns = Bloomberg ticker strings, or renamed columns when tickers
        is a dict).

    Raises
    ------
    ImportError
        If pdblp is not installed.
    pxtsValidationError
        If the returned DataFrame does not have a DatetimeIndex.
    """
    try:
        import pdblp
    except ImportError:
        raise ImportError(
            "pdblp required for read_bdh(). Install with: pip install pdblp and https://www.bloomberg.com/professional/support/api-library/"
        )

    if isinstance(tickers, dict):
        rename_map = tickers  # {new_name: bloomberg_ticker}
        ticker_list = list(tickers.values())
    else:
        rename_map = None
        if isinstance(tickers, str):
            tickers = [tickers]
        ticker_list = tickers

    start_date = pd.to_datetime(start).strftime("%Y%m%d")
    if end is None:
        end_date = pd.Timestamp.today().strftime("%Y%m%d")
    else:
        end_date = pd.to_datetime(end).strftime("%Y%m%d")

    # Establish a connection
    con = pdblp.BCon(port=8194, timeout=5000)
    con.start()

    # Get historical data for securities and fields
    try:
        raw = con.bdh(ticker_list, field, start_date, end_date)
    finally:
        con.stop()

    data = raw.loc[:, ticker_list]
    data.columns = ticker_list
    data.index = pd.to_datetime(data.index)

    if rename_map is not None:
        # rename_map is {new_name: bloomberg_ticker}; invert for DataFrame.rename
        data = data.rename(columns={v: k for k, v in rename_map.items()})

    return validate_ts(data)


def read_mb(series) -> pd.DataFrame:
    """Fetch Macrobond historical time series data.

    Calls mda.get_series, normalises the response, and returns a validated
    wide-format DataFrame with a DatetimeIndex and one column per series.
    The full available history is returned; slice the result as needed.

    Requires macrobond_data_api (optional dependency). Install with:
        pip install macrobond-data-api

    Parameters
    ----------
    series : list of str, str, or dict
        Macrobond series identifiers, e.g. ['hg_c1_cl', 'hg_c2_cl'].
        If a dict is provided, keys are the desired output column names and
        values are the Macrobond series identifiers, e.g.
        {'Copper 1M': 'hg_c1_cl', 'Copper 2M': 'hg_c2_cl'}.

    Returns
    -------
    pd.DataFrame
        Wide-format DataFrame with DatetimeIndex (rows = dates, columns =
        FullDescription from Macrobond metadata, or dict keys when series is
        a dict). Series with disjoint date ranges are aligned on the union of
        all dates; missing observations are NaN.

    Raises
    ------
    ImportError
        If macrobond_data_api is not installed.
    ValueError
        If any requested series returns an error from the Macrobond API.
    RuntimeError
        If the get_series call itself raises an unexpected exception.
    pxtsValidationError
        If the returned DataFrame does not have a DatetimeIndex.
    """
    try:
        import macrobond_data_api as mda
    except ImportError:
        raise ImportError(
            "macrobond_data_api required for read_mb(). "
            "Install with: pip install macrobond-data-api"
        )

    if isinstance(series, dict):
        rename_map = series  # {new_name: macrobond_id}
        series_list = list(series.values())
    else:
        rename_map = None
        if isinstance(series, str):
            series = [series]
        series_list = list(series)

    # Fetch entities from Macrobond (notebook cell 1)
    try:
        entities = mda.get_series(series_list)
    except Exception as exc:
        raise RuntimeError(f"macrobond_data_api.get_series failed: {exc}") from exc

    # Defensive error check — get_series may return error entities rather than raising
    for entity in entities:
        if getattr(entity, "is_error", False):
            name = getattr(entity, "name", "(unknown)")
            msg = getattr(entity, "error_message", str(entity))
            raise ValueError(f"Macrobond series error — '{name}': {msg}")

    # Normalise to a flat DataFrame (notebook cell 2)
    df_raw = pd.json_normalize([x.to_dict() for x in entities])

    # Explode Dates and Values into long format (notebook cell 3)
    df_long = (
        df_raw.set_index(["metadata.PrimName", "metadata.FullDescription"])[
            ["Dates", "Values"]
        ]
        .apply(pd.Series.explode)
        .reset_index()
    )

    # PrimName → FullDescription mapping used for column ordering
    prim_to_full = dict(
        zip(df_raw["metadata.PrimName"], df_raw["metadata.FullDescription"])
    )

    # Pivot to wide format; mismatched date ranges produce NaN automatically
    df_wide = df_long.pivot(
        index="Dates", columns="metadata.FullDescription", values="Values"
    )
    df_wide.index = pd.to_datetime(df_wide.index)
    df_wide.index.name = "date"
    df_wide.columns.name = None
    df_wide = df_wide.astype(float)

    if rename_map is not None:
        full_to_new = {prim_to_full[v]: k for k, v in rename_map.items()}
        df_wide = df_wide.rename(columns=full_to_new)
        df_wide = df_wide[list(rename_map.keys())]
    else:
        ordered_cols = [prim_to_full[p] for p in series_list]
        df_wide = df_wide[ordered_cols]

    return validate_ts(df_wide)


def read_xlsx(
    path: Union[str, Path],
    *,
    tz: str | None = None,
    date_format: str | None = None,
    datetime_col: str = "A",
    colname_row: int = 1,
    values_ref: str | None = None,
    sheet: str | None = None,
) -> pd.DataFrame:
    """Read an Excel (.xlsx or .xlsm) file into a DataFrame with a DatetimeIndex.

    One column supplies the datetime index; one row supplies column headers.
    Columns whose header cell is empty or whitespace-only are silently skipped.

    When *values_ref* is omitted the data extent is auto-detected:
    - Columns: scan *colname_row* rightward from the column after *datetime_col*;
      the range ends at the last non-empty header cell (empty cells in between
      are skipped, not treated as terminators).
    - Rows: scan *datetime_col* downward from *colname_row* + 1; stop at the
      first empty cell.

    When *values_ref* is supplied (e.g. ``'B2:BF2854'``) those bounds drive
    which rows and columns are read for data values. *datetime_col* and
    *colname_row* are still used to pull the index values and column names
    respectively.

    Parameters
    ----------
    path : str or Path
        Path to the Excel file. ``.xlsx`` and ``.xlsm`` are accepted;
        ``.xls`` raises ``ValueError``.
    tz : str or None
        Timezone to localize (or convert) the index to, e.g. ``'US/Eastern'``.
    date_format : str or None
        Explicit strptime format string (e.g. ``'%Y-%m-%d'``). If ``None``,
        auto-detected from the first value in *datetime_col* when dates are
        stored as strings; ignored when openpyxl returns native datetime objects.
    datetime_col : str
        Excel column letter(s) that contains the datetime index. Default ``'A'``.
    colname_row : int
        1-indexed row number containing the column headers. Default ``1``.
    values_ref : str or None
        Excel range for the data values, e.g. ``'B2:BF2854'``. Does not
        include the index column or header row — those come from *datetime_col*
        and *colname_row*. If ``None``, the range is auto-detected.
    sheet : str or None
        Sheet name to read. If ``None``, the first sheet is used.

    Returns
    -------
    pd.DataFrame
        DataFrame with a DatetimeIndex.

    Raises
    ------
    ValueError
        If the file has a ``.xls`` extension, or if no data is found.
    ImportError
        If openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.utils import column_index_from_string
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        raise ImportError(
            "openpyxl is required for read_xlsx(). "
            "Install with: pip install pxts[excel]"
        )

    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".xls":
        raise ValueError(
            "read_xlsx() does not support legacy .xls files. "
            "Save the file as .xlsx or .xlsm first."
        )
    if suffix not in (".xlsx", ".xlsm"):
        raise ValueError(
            f"read_xlsx() expects a .xlsx or .xlsm file, got '{suffix}'."
        )

    wb = openpyxl.load_workbook(path, data_only=True)

    ws = wb.worksheets[0] if sheet is None else wb[sheet]

    datetime_col_idx = column_index_from_string(datetime_col.upper())

    # ------------------------------------------------------------------
    # Determine the values range
    # ------------------------------------------------------------------
    if values_ref is not None:
        min_col, data_start_row, max_col, data_end_row = range_boundaries(
            values_ref.upper()
        )
    else:
        # Scan colname_row rightward to find the extent of non-empty headers.
        first_data_col: int | None = None
        last_data_col: int | None = None
        for col_idx in range(datetime_col_idx + 1, (ws.max_column or 0) + 1):
            val = ws.cell(row=colname_row, column=col_idx).value
            if val is not None and str(val).strip():
                if first_data_col is None:
                    first_data_col = col_idx
                last_data_col = col_idx

        if first_data_col is None:
            raise ValueError(
                f"read_xlsx(): no column headers found in row {colname_row} "
                f"beyond column '{datetime_col}'."
            )

        min_col = first_data_col
        max_col = last_data_col  # type: ignore[assignment]

        # Scan datetime_col downward from colname_row+1; stop at first empty.
        data_start_row = colname_row + 1
        data_end_row = colname_row  # sentinel — updated in the loop
        for row_idx in range(data_start_row, (ws.max_row or 0) + 1):
            val = ws.cell(row=row_idx, column=datetime_col_idx).value
            if val is None or (isinstance(val, str) and not val.strip()):
                break
            data_end_row = row_idx

        if data_end_row < data_start_row:
            raise ValueError(
                f"read_xlsx(): no data rows found in column '{datetime_col}' "
                f"starting from row {data_start_row}."
            )

    # ------------------------------------------------------------------
    # Build column-name map, skipping empty / whitespace-only headers
    # ------------------------------------------------------------------
    col_map: dict[int, str] = {}
    for col_idx in range(min_col, max_col + 1):
        val = ws.cell(row=colname_row, column=col_idx).value
        if val is not None and str(val).strip():
            col_map[col_idx] = str(val).strip()

    if not col_map:
        raise ValueError(
            "read_xlsx(): all column headers in the values range are empty."
        )

    # Preserve the name of the index column (mirrors read_ts behaviour).
    _index_name_val = ws.cell(row=colname_row, column=datetime_col_idx).value
    index_name = str(_index_name_val).strip() if _index_name_val is not None else None

    col_indices = list(col_map.keys())
    col_names = list(col_map.values())

    # ------------------------------------------------------------------
    # Read rows: datetime values + data values
    # ------------------------------------------------------------------
    dt_raw: list = []
    rows_data: list[list] = []

    for row_idx in range(data_start_row, data_end_row + 1):
        dt_cell = ws.cell(row=row_idx, column=datetime_col_idx).value
        if dt_cell is None or (isinstance(dt_cell, str) and not dt_cell.strip()):
            continue
        dt_raw.append(dt_cell)
        rows_data.append(
            [ws.cell(row=row_idx, column=c).value for c in col_indices]
        )

    if not dt_raw:
        raise ValueError("read_xlsx(): no data rows found.")

    # ------------------------------------------------------------------
    # Parse datetime index
    # ------------------------------------------------------------------
    first_val = dt_raw[0]

    if date_format is not None:
        str_dates = [str(v) for v in dt_raw]
        if date_format == "ISO8601":
            index = pd.to_datetime(str_dates, format="ISO8601")
        else:
            index = pd.to_datetime(str_dates, format=date_format)
    elif isinstance(first_val, (datetime.datetime, datetime.date)):
        # openpyxl already parsed date-formatted cells into Python objects.
        index = pd.to_datetime(dt_raw)
    elif isinstance(first_val, (int, float)):
        # Excel serial date numbers (rare with data_only=True, but possible).
        converted = [
            datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(v))
            for v in dt_raw
        ]
        index = pd.to_datetime(converted)
    else:
        # String dates — reuse the same auto-detection logic as read_ts.
        str_dates = [str(v) for v in dt_raw]
        detected_format, dayfirst = _detect_date_format(str_dates[0])
        if detected_format == "ISO8601":
            index = pd.to_datetime(str_dates, format="ISO8601")
        elif detected_format == "mixed":
            index = pd.to_datetime(str_dates, format="mixed", dayfirst=dayfirst)
        else:
            index = pd.to_datetime(str_dates, format=detected_format, dayfirst=dayfirst)

    index.name = index_name

    # ------------------------------------------------------------------
    # Build and return DataFrame
    # ------------------------------------------------------------------
    df = pd.DataFrame(rows_data, index=index, columns=col_names)

    if tz is not None:
        if df.index.tz is None:
            df.index = df.index.tz_localize(tz)
        else:
            df.index = df.index.tz_convert(tz)

    return validate_ts(df)